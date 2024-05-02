import openpyxl

def unmerge_fill(file_path):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active  # 获取活动表

    # 收集所有合并单元格的范围到一个列表中
    merged_cells_ranges = list(worksheet.merged_cells.ranges)

    # 遍历所有先前收集的合并单元格范围
    for merge_cell in merged_cells_ranges:
        min_col, min_row, max_col, max_row = merge_cell.bounds
        top_left_cell_value = worksheet.cell(row=min_row, column=min_col).value

        # 取消合并单元格
        worksheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

        # 将原始值复制到所有相关单元格
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                worksheet.cell(row=row, column=col, value=top_left_cell_value)

    # 保存工作簿
    workbook.save("unmerged_" + file_path)

# 调用函数
unmerge_fill("card.xlsx")
